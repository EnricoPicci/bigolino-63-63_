#!/usr/bin/env python3
"""
Process the invitation Excel file and produce a status report.

Reads "invitation-lists/Lista della festa 63-63.xlsx" and produces
"invitation-lists/invitation-status.xlsx" with three sheets:
  1. Email Sent       - people already emailed, with names when known
  2. Email NOT Sent   - people with emails who haven't been emailed yet
  3. Other Emails     - emails from other sheets not in either list above
"""

import os
import re
import unicodedata
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
INPUT_FILE = os.path.join(PROJECT_DIR, "invitation-lists", "Lista della festa 63-63.xlsx")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "invitation-lists", "invitation-status.xlsx")


def clean_email(raw):
    """Extract and normalize an email address from a cell value."""
    if raw is None:
        return None
    s = str(raw)
    # Remove unicode directional markers
    s = s.replace('\u202a', '').replace('\u202c', '')
    # Remove other control characters
    s = ''.join(c for c in s if unicodedata.category(c)[0] != 'C')
    s = s.strip()
    if not s or '@' not in s:
        return None
    # Extract email from "Name <email>" format
    match = re.search(r'<([^>]+@[^>]+)>', s)
    if match:
        s = match.group(1).strip()
    # Remove any remaining whitespace
    s = s.strip()
    # Basic validation
    if re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', s):
        return s
    return None


def extract_emails_from_cells(row, columns):
    """Extract emails from specific columns in a row."""
    emails = []
    for col_letter in columns:
        col_idx = ord(col_letter.upper()) - ord('A')
        if col_idx < len(row):
            email = clean_email(row[col_idx])
            if email:
                emails.append(email)
    return emails


def read_named_sheet(ws, email_columns, sheet_label):
    """
    Read a sheet with named invitees (cols A=last name, B=first name).
    Last names propagate down for family groups.
    Returns list of (last_name, first_name, email, sheet_label).
    """
    results = []
    current_last_name = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
        if row_idx == 0:
            continue  # skip header

        col_a = str(row[0]).strip() if row[0] else ""
        col_b = str(row[1]).strip() if row[1] and len(row) > 1 else ""

        # Update family last name when column A has a value
        if col_a and col_a not in ('', ' '):
            current_last_name = col_a

        first_name = col_b

        if not first_name or first_name in ('', ' '):
            # No person on this row, but might still have an email
            pass

        emails = extract_emails_from_cells(row, email_columns)
        for email in emails:
            results.append((current_last_name, first_name, email, sheet_label))

    return results


def main():
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    # ── Step 1: Collect SENT emails from actual mail recipient list ──
    # This is the source of truth — the actual list of recipients of the first mail.
    ACTUAL_MAIL_RECIPIENTS = [
        "abigi@remedia.it",
        "aconfet@thoughtworks.com",
        "adallazuanna@hotmail.com",
        "Alberto.saccardi@nunatac.it",
        "aless.camon@gmail.com",
        "alexpicci@libero.it",
        "andrea.caligaris@generali.com",
        "andrea.cartamantiglia@beplex.com",
        "andrea.manzitti@beplex.com",
        "angelo@fondall.com",
        "antonella.andrioli63@gmail.com",
        "avv.bertin@francescodibartolo.it",
        "avv.carlobertin@libero.it",
        "avv.mcali@gmail.com",
        "benedettasdn@gmail.com",
        "borgheri@iuav.it",
        "c.sainati@dedalosnc.it",
        "carlo@dalla-libera.com",
        "castiglionimauro60@gmail.com",
        "checco.magro63@gmail.com",
        "claudia.visentin@gmail.com",
        "consolata.morassutti@gmail.com",
        "Cristina.Mezzabarba@beplex.com",
        "danielaserfilippi@yahoo.it",
        "dantedda01@gmail.com",
        "davide.ragno@virgilio.it",
        "dflpadova@gmail.com",
        "donatella.picarelli@gmail.com",
        "drmarcocomaschi@gmail.com",
        "enricorabbioso@gmail.com",
        "equilibrioinstabile@tiscali.it",
        "esanti00@engr.sc.edu",
        "f.varotto@marsilioeditori.it",
        "flavio.addolorato@intesasanpaolo.com",
        "fonorato@gmail.com",
        "gbn.pdv@gmail.com",
        "Gbrcarrera@gmail.com",
        "ginalu@alice.it",
        "giovannabilato@hotmail.it",
        "giovanni.guglielmetti@beplex.com",
        # "giulialion@alice.it",  # removed: not confirmed sent
        "flavio.addolorato@outlook.com",  # alternate address, person was emailed
        "sabatomanzo@gmail.com",  # alternate address, person was emailed
        "giulialion63@gmail.com",
        "grazialateiera@gmail.com",
        "gsoloni@gmail.com",
        "ilpisto@gmail.com",
        "info@pietrogorgone.it",
        "info@renzofeltrin.it",
        "isabellapansera@bananas.it",
        "laura61cerutti@gmail.com",
        "laurence.khoury@gmail.com",
        "lerrichetti@gmail.com",
        "lorenzo.negrato@ne-xt.org",
        "ltieghi@tampieri.com",
        "lucacavestro25@gmail.com",
        "lucacdo@gmail.com",
        "Marco.andry@gmail.com",
        "marcobertin.mb@gmail.com",
        "mariopiccinin@gmail.com",
        "mastrosimone.andrea@gmail.com",
        "Matteo.bonati@libero.it",
        "mdelfra@yahoo.it",
        "mgmissaggia@gmail.com",
        "michele.ulivi@msn.com",
        "micheleulivi@msn.com",
        "msalvini100@gmail.com",
        "Paola_pea@yahoo.it",
        "paola.ranzolin@gmail.com",
        "paolabertin.pb@gmail.com",
        "paolodallacosta@gmail.com",
        "pferrari@remedia.it",
        "r.semenzato@gmail.com",
        "rachel.erdman@gmail.com",
        "raffaello.graziotto@gmail.com",
        "romoletta1@yahoo.it",
        "rossella.franchi@gmail.com",
        "s.manzo@mhore.it",
        "sartore.marco@gmail.com",
        "silviagirardi@hotmail.it",
        "simonetta.candela@cliffordchance.com",
        "simonettaandrioli@virgilio.it",
        "SPetruzziello@disaronno.it",
        "tommaso.riccoboni@gmail.com",
        "toni.fregnan@gmail.com",
        "vittolor@yahoo.it",
        "ziliottomara@gmail.com",
        "zingalesfrancesca66@gmail.com",
        "Zocky@libero.it",
        "mirella.cerutti@gmail.com",
    ]
    sent_emails = {e.lower() for e in ACTUAL_MAIL_RECIPIENTS}

    # Also collect non-B-column emails from email-sent-63 as "not yet sent"
    ws_sent = wb["email-sent-63"]
    not_yet_sent_from_sent_sheet = set()
    for row_idx, row in enumerate(ws_sent.iter_rows(min_row=1, values_only=True)):
        if row_idx == 0:
            continue
        for col_idx, cell in enumerate(row):
            email = clean_email(cell)
            if email:
                col_letter = get_column_letter(col_idx + 1)
                if col_letter not in ('B', 'I') and email.lower() not in sent_emails:
                    not_yet_sent_from_sent_sheet.add(email.lower())

    print(f"Emails sent (actual mail): {len(sent_emails)}")
    print(f"Not yet sent (email-sent-63 other cols): {len(not_yet_sent_from_sent_sheet)}")

    # ── Step 2: Collect named invitees from all named sheets ──
    # Each sheet has email in different columns
    sheet_configs = [
        ("Invitati_63", ['H', 'I', 'J', 'N'], "Invitati_63"),
        ("Enrico",      ['H'],                 "Enrico"),
        ("Gigio",       ['D'],                 "Gigio"),
        ("Giovanni",    ['G'],                 "Giovanni"),
        ("Stefano",     ['D'],                 "Stefano"),
        ("Invitati_60", ['H', 'I', 'J', 'N'], "Invitati_60"),
    ]

    # email -> {last_name, first_name, sources: set of sheet names}
    email_to_person = {}

    for sheet_name, email_cols, label in sheet_configs:
        ws = wb[sheet_name]
        # Also check "Mail fallita" columns
        fail_cols = []
        if sheet_name == "Enrico":
            fail_cols = ['J']
        elif sheet_name == "Stefano":
            fail_cols = ['I']

        entries = read_named_sheet(ws, email_cols + fail_cols, label)
        for last_name, first_name, email, src in entries:
            key = email.lower()
            if key not in email_to_person:
                email_to_person[key] = {
                    'last_name': last_name,
                    'first_name': first_name,
                    'email': email,  # preserve original case
                    'sources': set(),
                }
            email_to_person[key]['sources'].add(src)
            # Update name if we have a better one
            if last_name and not email_to_person[key]['last_name']:
                email_to_person[key]['last_name'] = last_name
            if first_name and not email_to_person[key]['first_name']:
                email_to_person[key]['first_name'] = first_name

    # Add emails from non-B columns of email-sent-63 that aren't already known
    for email_lower in not_yet_sent_from_sent_sheet:
        if email_lower not in email_to_person:
            email_to_person[email_lower] = {
                'last_name': '',
                'first_name': '',
                'email': email_lower,
                'sources': {'email-sent-63 (not yet sent)'},
            }

    print(f"Named invitees with emails: {len(email_to_person)}")

    # ── Step 3: Build the three output lists ──

    # List 1: Emails sent — with names if known
    list_sent = []
    for email_lower in sorted(sent_emails):
        person = email_to_person.get(email_lower, {})
        last_name = person.get('last_name', '')
        first_name = person.get('first_name', '')
        original_email = person.get('email', email_lower)
        sources = person.get('sources', set())
        # Determine which organizer's list they're on
        organizer_lists = sorted(sources & {'Enrico', 'Gigio', 'Giovanni', 'Stefano'})
        list_sent.append({
            'last_name': last_name,
            'first_name': first_name,
            'email': original_email,
            'lists': ', '.join(organizer_lists),
        })

    # List 2: Emails NOT sent — people with emails who haven't been emailed
    list_not_sent = []
    for email_lower, person in sorted(email_to_person.items(), key=lambda x: x[0]):
        if email_lower not in sent_emails:
            organizer_lists = sorted(person['sources'] & {'Enrico', 'Gigio', 'Giovanni', 'Stefano'})
            all_sources = sorted(person['sources'])
            list_not_sent.append({
                'last_name': person['last_name'],
                'first_name': person['first_name'],
                'email': person['email'],
                'organizer_lists': ', '.join(organizer_lists),
                'all_sources': ', '.join(all_sources),
            })

    # List 3: Other emails — from all other sheets, not in sent or named lists
    all_known_emails = sent_emails | set(email_to_person.keys())
    other_sheets = ["email sent", "list of email sent", "tutte emails", "mail onlus"]
    other_emails = {}  # email -> set of source sheets
    for sheet_name in other_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, values_only=True):
            for cell in row:
                email = clean_email(cell)
                if email and email.lower() not in all_known_emails:
                    key = email.lower()
                    if key not in other_emails:
                        other_emails[key] = {'email': email, 'sources': set()}
                    other_emails[key]['sources'].add(sheet_name)

    list_other = []
    for email_lower in sorted(other_emails.keys()):
        info = other_emails[email_lower]
        list_other.append({
            'email': info['email'],
            'sources': ', '.join(sorted(info['sources'])),
        })

    print(f"\nResults:")
    print(f"  Email Sent:     {len(list_sent)}")
    print(f"  Email NOT Sent: {len(list_not_sent)}")
    print(f"  Other Emails:   {len(list_other)}")

    # ── Step 4: Write output Excel ──
    out_wb = openpyxl.Workbook()

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4a7c59", end_color="4a7c59", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_sheet(ws, headers, col_widths):
        """Apply headers and styling to a worksheet."""
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def gmail_format(first_name, last_name, email):
        """Build 'First Last <email>' avoiding duplicate name parts."""
        # Some sheets put the full name in the last_name column (e.g. "Andrea Agressi")
        # If last_name already contains first_name, use last_name as-is
        if last_name and first_name and first_name.lower() in last_name.lower():
            name = last_name
        else:
            name = ' '.join(p for p in [first_name, last_name] if p)
        return f"{name} <{email}>" if name else email

    # Sheet 1: Email Sent
    ws1 = out_wb.active
    ws1.title = "Email Sent"
    style_sheet(ws1,
        ["Last Name", "First Name", "Email", "Gmail Format", "Organizer List"],
        [18, 15, 35, 45, 15])
    for item in list_sent:
        ws1.append([
            item['last_name'],
            item['first_name'],
            item['email'],
            gmail_format(item['first_name'], item['last_name'], item['email']),
            item['lists'],
        ])

    # Sheet 2: Email NOT Sent
    ws2 = out_wb.create_sheet("Email NOT Sent")
    style_sheet(ws2,
        ["Last Name", "First Name", "Email", "Gmail Format", "Organizer List", "Found In Sheets"],
        [18, 15, 35, 45, 18, 30])
    for item in list_not_sent:
        ws2.append([
            item['last_name'],
            item['first_name'],
            item['email'],
            gmail_format(item['first_name'], item['last_name'], item['email']),
            item['organizer_lists'],
            item['all_sources'],
        ])

    # Sheet 3: Other Emails
    ws3 = out_wb.create_sheet("Other Emails")
    style_sheet(ws3,
        ["Email", "Found In Sheets"],
        [35, 40])
    for item in list_other:
        ws3.append([item['email'], item['sources']])

    # Apply border to all data rows
    for ws in [ws1, ws2, ws3]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    out_wb.save(OUTPUT_FILE)
    print(f"\nOutput written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
